"""Excel .xlsx 文档解析器。"""

from __future__ import annotations

from pathlib import Path
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from day04_app.common.exceptions import BusinessException
from day04_app.schemas.knowledge_schema import ParsedDocument, ParsedDocumentSegment


class ExcelDocumentParser:
    """按工作表和行读取 Excel，并保留精确到单元格坐标的来源信息。"""

    supported_file_types = frozenset({"xlsx"})

    def parse(self, *, document_id: str, file_path: Path) -> ParsedDocument:
        try:
            # read_only 降低大表解析的内存占用；data_only=False 保留公式本身而非不可靠的缓存值。
            workbook = load_workbook(file_path, read_only=True, data_only=False, keep_links=False)
        except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile) as exc:
            raise BusinessException(code=40064, message="Excel 文件无法读取或文件已损坏") from exc

        segments: list[ParsedDocumentSegment] = []
        non_empty_row_count = 0
        sheet_names = list(workbook.sheetnames)
        try:
            for worksheet in workbook.worksheets:
                for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                    cell_texts: list[str] = []
                    contains_formula = False
                    for cell in row:
                        if cell.value is None or str(cell.value).strip() == "":
                            continue
                        cell_value = str(cell.value).strip()
                        contains_formula = contains_formula or cell.data_type == "f"
                        # 使用 A1 坐标而不是猜测表头，能兼容标题行、合并单元格和不规则工作表。
                        cell_texts.append(f"{cell.coordinate}: {cell_value}")
                    if not cell_texts:
                        continue
                    non_empty_row_count += 1
                    segments.append(
                        ParsedDocumentSegment(
                            segment_index=len(segments),
                            text=" | ".join(cell_texts),
                            location=f"Sheet:{worksheet.title}/Row:{row_index}",
                            metadata={
                                "source_type": "sheet_row",
                                "sheet_name": worksheet.title,
                                "row_index": row_index,
                                "contains_formula": contains_formula,
                            },
                        )
                    )
        except (OSError, ValueError, zipfile.BadZipFile) as exc:
            raise BusinessException(code=40064, message="Excel 内容读取失败") from exc
        finally:
            workbook.close()

        if not segments:
            raise BusinessException(code=40065, message="Excel 中没有可用于知识库的非空单元格")

        return ParsedDocument(
            document_id=document_id,
            file_name=file_path.name,
            file_type="xlsx",
            parser_name=self.__class__.__name__,
            segments=segments,
            metadata={
                "sheet_names": sheet_names,
                "sheet_count": len(sheet_names),
                "non_empty_row_count": non_empty_row_count,
                "segment_count": len(segments),
            },
        )
